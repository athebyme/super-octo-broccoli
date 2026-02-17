# -*- coding: utf-8 -*-
"""
Универсальный модуль экспорта данных

Поддерживает форматы: CSV, XLSX, текст (через запятую / разделитель).
Позволяет выбирать произвольные колонки для выгрузки.
"""
import csv
import logging
from io import BytesIO, StringIO
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence

from flask import make_response

logger = logging.getLogger('data_export')


# ==================== ОПРЕДЕЛЕНИЯ КОЛОНОК ====================

# Доступные колонки для карточек товаров (Product)
PRODUCT_COLUMNS = {
    'nm_id': {'label': 'Артикул WB', 'getter': lambda p: p.get('nmId') or p.get('nm_id', '')},
    'vendor_code': {'label': 'Артикул продавца', 'getter': lambda p: p.get('vendorCode') or p.get('vendor_code', '')},
    'title': {'label': 'Наименование', 'getter': lambda p: p.get('title', '')},
    'brand': {'label': 'Бренд', 'getter': lambda p: p.get('brand', '')},
    'category': {'label': 'Категория', 'getter': lambda p: p.get('object_name') or p.get('subjectName', '')},
    'price': {'label': 'Цена', 'getter': lambda p: p.get('price', '')},
    'barcode': {'label': 'Баркод', 'getter': lambda p: _extract_barcode(p)},
}

# Колонки для заблокированных карточек
BLOCKED_CARD_COLUMNS = {
    'nm_id': {'label': 'Артикул WB', 'getter': lambda p: p.get('nmId', '')},
    'vendor_code': {'label': 'Артикул продавца', 'getter': lambda p: p.get('vendorCode', '')},
    'title': {'label': 'Наименование', 'getter': lambda p: p.get('title', '')},
    'brand': {'label': 'Бренд', 'getter': lambda p: p.get('brand', '')},
    'reason': {'label': 'Причина блокировки', 'getter': lambda p: p.get('reason', '')},
}

# Колонки для скрытых карточек
SHADOWED_CARD_COLUMNS = {
    'nm_id': {'label': 'Артикул WB', 'getter': lambda p: p.get('nmId', '')},
    'vendor_code': {'label': 'Артикул продавца', 'getter': lambda p: p.get('vendorCode', '')},
    'title': {'label': 'Наименование', 'getter': lambda p: p.get('title', '')},
    'brand': {'label': 'Бренд', 'getter': lambda p: p.get('brand', '')},
    'nm_rating': {'label': 'Рейтинг', 'getter': lambda p: p.get('nmRating', '')},
}

# Реестр всех наборов колонок
COLUMN_SETS = {
    'products': PRODUCT_COLUMNS,
    'blocked': BLOCKED_CARD_COLUMNS,
    'shadowed': SHADOWED_CARD_COLUMNS,
}


def _extract_barcode(item: dict) -> str:
    """Извлечь первый баркод из карточки товара"""
    sizes = item.get('sizes', [])
    if sizes:
        skus = sizes[0].get('skus', [])
        if skus:
            return skus[0]
    return ''


# ==================== ЭКСПОРТ ====================

def export_data(
    data: List[Dict[str, Any]],
    columns: List[str],
    column_defs: Dict[str, dict],
    fmt: str = 'csv',
    filename_prefix: str = 'export',
    separator: str = ', ',
    single_column_for_text: Optional[str] = None,
):
    """
    Универсальная функция экспорта данных в разные форматы.

    Args:
        data: Список словарей с данными
        columns: Список ключей колонок для выгрузки (порядок сохраняется)
        column_defs: Определения колонок (из COLUMN_SETS или кастомные)
        fmt: Формат экспорта ('csv', 'xlsx', 'text')
        filename_prefix: Префикс имени файла
        separator: Разделитель для текстового формата
        single_column_for_text: Если задано, в текстовом формате выгружается
            только эта колонка (без заголовков). Удобно для списка артикулов.

    Returns:
        Flask Response с файлом для скачивания
    """
    # Валидация колонок
    valid_columns = [c for c in columns if c in column_defs]
    if not valid_columns:
        valid_columns = list(column_defs.keys())

    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')

    if fmt == 'xlsx':
        return _export_xlsx(data, valid_columns, column_defs, filename_prefix, timestamp)
    elif fmt == 'text':
        return _export_text(
            data, valid_columns, column_defs, filename_prefix, timestamp,
            separator, single_column_for_text
        )
    else:
        return _export_csv(data, valid_columns, column_defs, filename_prefix, timestamp)


def _export_csv(
    data: List[Dict[str, Any]],
    columns: List[str],
    column_defs: Dict[str, dict],
    filename_prefix: str,
    timestamp: str,
) -> Any:
    """Экспорт в CSV"""
    output = StringIO()
    output.write('\ufeff')  # UTF-8 BOM для Excel

    writer = csv.writer(output)

    # Заголовки
    headers = [column_defs[c]['label'] for c in columns]
    writer.writerow(headers)

    # Данные
    for item in data:
        row = [column_defs[c]['getter'](item) for c in columns]
        writer.writerow(row)

    csv_data = output.getvalue().encode('utf-8-sig')
    response = make_response(csv_data)
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = (
        f'attachment; filename={filename_prefix}_{timestamp}.csv'
    )
    return response


def _export_xlsx(
    data: List[Dict[str, Any]],
    columns: List[str],
    column_defs: Dict[str, dict],
    filename_prefix: str,
    timestamp: str,
) -> Any:
    """Экспорт в XLSX"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Данные'

    # Стили заголовков
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Заголовки
    headers = [column_defs[c]['label'] for c in columns]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Данные
    data_alignment = Alignment(vertical='center', wrap_text=True)
    for row_idx, item in enumerate(data, 2):
        for col_idx, col_key in enumerate(columns, 1):
            value = column_defs[col_key]['getter'](item)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = data_alignment

    # Автоширина колонок
    for col_idx, col_key in enumerate(columns, 1):
        max_length = len(column_defs[col_key]['label'])
        for row_idx in range(2, min(len(data) + 2, 102)):  # Проверяем первые 100 строк
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 4, 60)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

    # Закрепить первую строку
    ws.freeze_panes = 'A2'

    # Автофильтр
    if data:
        ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(columns))}{len(data) + 1}'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = (
        f'attachment; filename={filename_prefix}_{timestamp}.xlsx'
    )
    return response


def _export_text(
    data: List[Dict[str, Any]],
    columns: List[str],
    column_defs: Dict[str, dict],
    filename_prefix: str,
    timestamp: str,
    separator: str = ', ',
    single_column: Optional[str] = None,
) -> Any:
    """Экспорт в текстовый формат (значения через разделитель)"""
    if single_column and single_column in column_defs:
        # Режим одной колонки — просто значения через разделитель
        values = []
        for item in data:
            val = str(column_defs[single_column]['getter'](item))
            if val:
                values.append(val)
        text_data = separator.join(values)
    else:
        # Все выбранные колонки, каждая строка данных на отдельной строке
        lines = []
        # Заголовок
        headers = [column_defs[c]['label'] for c in columns]
        lines.append(separator.join(headers))
        # Данные
        for item in data:
            row_values = [str(column_defs[c]['getter'](item)) for c in columns]
            lines.append(separator.join(row_values))
        text_data = '\n'.join(lines)

    response = make_response(text_data.encode('utf-8'))
    response.headers['Content-Type'] = 'text/plain; charset=utf-8'
    response.headers['Content-Disposition'] = (
        f'attachment; filename={filename_prefix}_{timestamp}.txt'
    )
    return response


def get_available_columns(column_set: str) -> Dict[str, str]:
    """
    Получить доступные колонки для набора данных.

    Args:
        column_set: Имя набора ('products', 'blocked', 'shadowed')

    Returns:
        Словарь {ключ_колонки: человекочитаемое_название}
    """
    defs = COLUMN_SETS.get(column_set, {})
    return {key: val['label'] for key, val in defs.items()}
